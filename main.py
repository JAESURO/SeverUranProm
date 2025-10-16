import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from config import config
import os
from datetime import datetime, timedelta
import numpy as np


def get_public_tables(cursor):
    cursor.execute(
        """
        select table_name
        from information_schema.tables
        where table_schema = 'public' and table_type = 'BASE TABLE'
        order by table_name
        """
    )
    return [row[0] for row in cursor.fetchall()]


def get_columns(cursor, table_name):
    cursor.execute(
        """
        select column_name, data_type
        from information_schema.columns
        where table_schema = 'public' and table_name = %s
        order by ordinal_position
        """,
        (table_name,),
    )
    return cursor.fetchall()


def get_foreign_key_joins(cursor):
    cursor.execute(
        """
        SELECT
            tc.table_name       AS source_table,
            kcu.column_name     AS source_column,
            ccu.table_name      AS target_table,
            ccu.column_name     AS target_column
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
          ON tc.constraint_name = kcu.constraint_name
         AND tc.table_schema = kcu.table_schema
        JOIN information_schema.constraint_column_usage ccu
          ON ccu.constraint_name = tc.constraint_name
         AND ccu.table_schema = tc.table_schema
        WHERE tc.constraint_type = 'FOREIGN KEY'
          AND tc.table_schema = 'public'
        ORDER BY source_table, target_table
        """
    )
    return cursor.fetchall()


def pick_first_numeric_column(columns):
    numeric_types = {
        'smallint', 'integer', 'bigint', 'real', 'double precision',
        'numeric', 'decimal'
    }
    for name, dtype in columns:
        if dtype in numeric_types:
            return name
    return None


def pick_first_text_column(columns):
    text_types = {'text', 'varchar', 'character varying', 'character', 'citext'}
    for name, dtype in columns:
        if dtype in text_types:
            return name
    return None


def export_to_excel(dataframes_dict, filename):
    """Export data to Excel with formatting"""
    filepath = f"exports/{filename}"
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            worksheet.freeze_panes = "B2"
            
            worksheet.auto_filter.ref = worksheet.dimensions
            
            for col in worksheet.iter_cols(min_row=2):
                if any(isinstance(cell.value, (int, float)) for cell in col[1:] if cell.value is not None):
                    col_letter = col[0].column_letter
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFAA0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                        end_type="max", end_color="FF00AA00"
                    )
                    worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(df)+1}", rule)
    
    total_rows = sum(len(df) for df in dataframes_dict.values())
    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")
    return filepath


def create_charts(df, chart_name, chart_type, business_question):
    plt.style.use('default')
    plt.rcParams['figure.figsize'] = (12, 8)
    plt.rcParams['font.size'] = 10
    
    chart_path = f"charts/{chart_name}_{chart_type}.png"
    
    if chart_type == 'pie':
        if len(df.columns) >= 2:
            labels = df.iloc[:, 0].astype(str)
            values = df.iloc[:, 1] if df.iloc[:, 1].dtype in ['int64', 'float64'] else range(len(df))
            plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
            plt.title(f'Pie Chart: {business_question}')
        
    elif chart_type == 'bar':
        if len(df.columns) >= 2:
            x_data = df.iloc[:, 0].astype(str)
            y_data = df.iloc[:, 1] if df.iloc[:, 1].dtype in ['int64', 'float64'] else range(len(df))
            plt.bar(x_data, y_data)
            plt.title(f'Bar Chart: {business_question}')
            plt.xticks(rotation=45)
            plt.xlabel(df.columns[0])
            plt.ylabel(df.columns[1])
        
    elif chart_type == 'barh':
        if len(df.columns) >= 2:
            y_data = df.iloc[:, 0].astype(str)
            x_data = df.iloc[:, 1] if df.iloc[:, 1].dtype in ['int64', 'float64'] else range(len(df))
            plt.barh(y_data, x_data)
            plt.title(f'Horizontal Bar Chart: {business_question}')
            plt.xlabel(df.columns[1])
            plt.ylabel(df.columns[0])
        
    elif chart_type == 'line':
        if len(df.columns) >= 2:
            text_cols = [c for c in df.columns if df[c].dtype not in ['int64', 'float64']]
            x_col = text_cols[0] if len(text_cols) > 0 else df.columns[0]
            num_cols = [c for c in df.columns if df[c].dtype in ['int64', 'float64']]
            y_col = num_cols[0] if len(num_cols) > 0 else df.columns[1]
            
            df_plot = df[[x_col, y_col]].copy()
            if df_plot[y_col].dtype == 'object':
                df_plot[y_col] = pd.to_numeric(df_plot[y_col], errors='coerce')
            df_plot = df_plot.dropna()
            
            if len(df_plot) > 0:
                x_data = df_plot[x_col].astype(str).tolist()
                y_data = df_plot[y_col].tolist()
                plt.plot(x_data, y_data, marker='o')
                plt.title(f'Line Chart: {business_question}')
                plt.xlabel(x_col)
                plt.ylabel(y_col)
                plt.xticks(rotation=45)
            else:
                print("Insufficient data for line chart")
        
    elif chart_type == 'hist':
        if len(df.columns) >= 1:
            numeric_col = None
            for col in df.columns:
                series = df[col]
                if series.dtype == 'object':
                    series = pd.to_numeric(series, errors='coerce')
                if series.dtype in ['int64', 'float64'] and not series.isna().all():
                    numeric_col = col
                    df[col] = series
                    break
            
            if numeric_col is not None:
                data = df[numeric_col].dropna()
                data = data[data > 0]
                if len(data) > 0:
                    bins = min(20, max(5, len(data)//4))
                    plt.hist(data, bins=bins, alpha=0.75, edgecolor='black', color='skyblue')
                    plt.title(f'Histogram: {business_question}')
                    plt.xlabel(numeric_col)
                    plt.ylabel('Frequency')
                    plt.grid(True, alpha=0.3)
                else:
                    print(f"No numeric data for histogram in column {numeric_col}")
            else:
                print("No suitable numeric column found for histogram")
        
    elif chart_type == 'scatter':
        if len(df.columns) >= 2:
            text_cols = [c for c in df.columns if df[c].dtype not in ['int64', 'float64']]
            num_cols = [c for c in df.columns if df[c].dtype in ['int64', 'float64']]
            
            x_col = text_cols[0] if len(text_cols) > 0 else df.columns[0]
            y_col = num_cols[0] if len(num_cols) > 0 else df.columns[1]
            
            df_plot = df[[x_col, y_col]].copy()
            if df_plot[y_col].dtype == 'object':
                df_plot[y_col] = pd.to_numeric(df_plot[y_col], errors='coerce')
            df_plot = df_plot.dropna()
            
            if len(df_plot) > 0:
                x_data = df_plot[x_col].astype(str).tolist()
                y_data = df_plot[y_col].tolist()
                plt.scatter(x_data, y_data, alpha=0.6)
                plt.title(f'Scatter Plot: {business_question}')
                plt.xlabel(x_col)
                plt.ylabel(y_col)
                plt.xticks(rotation=45)
            else:
                print("Insufficient data for scatter plot")
    
    plt.tight_layout()
    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Created chart: {chart_type} - {business_question}")
    print(f"Number of rows: {len(df)}")
    print(f"Saved to: {chart_path}")
    print("-" * 50)
    
    return chart_path


def create_interactive_chart(df, chart_name, business_question):
    df_work = df.copy()
    
    year_columns = [col for col in df_work.columns if col.startswith('Production_')]
    
    if len(year_columns) > 0:
        df_long = pd.melt(df_work, 
                         id_vars=['Country'], 
                         value_vars=year_columns,
                         var_name='Year', 
                         value_name='Production')
        
        df_long['Year'] = df_long['Year'].str.replace('Production_', '').astype(int)
        df_long = df_long.dropna(subset=['Production'])
        
        if len(df_long) > 0:
            df_long['Production'] = pd.to_numeric(df_long['Production'], errors='coerce')
            df_long = df_long.dropna(subset=['Production'])
            
            fig = px.scatter(df_long, 
                           x='Country', 
                           y='Production',
                           animation_frame='Year',
                           size='Production',
                           hover_data=['Country', 'Year', 'Production'],
                           title=f'Interactive Chart: {business_question}',
                           labels={'Country': 'Country', 'Production': 'Uranium Production (tonnes)'})
            
            fig.update_layout(
                title_x=0.5,
                xaxis_title='Country',
                yaxis_title='Uranium Production (tonnes)',
                width=1200,
                height=700,
                xaxis={'tickangle': 45}
            )
            
            fig.show()
            
            print(f"Created interactive chart: {business_question}")
            print(f"Data points: {len(df_long)}")
            print(f"Countries: {df_long['Country'].nunique()}")
            print(f"Years: {df_long['Year'].min()}-{df_long['Year'].max()}")
            print("Chart displayed in browser with time slider")
            print("-" * 50)
            
            return fig
        else:
            print("No data available for interactive chart after cleaning")
    else:
        print("No year columns found for time series visualization")
    
    return None


def run_queries():
    conn = None
    try:
        params = config()
        print('Connecting to the PostgreSQL database...')
        conn = psycopg2.connect(**params)
        cur = conn.cursor()

        cur.execute('select version()')
        print(cur.fetchone()[0])

        tables = get_public_tables(cur)
        if not tables:
            print('No user tables found in schema public. Exiting.')
            return

        visualization_queries = [
            {
                'name': '1. Nuclear power plants by country',
                'query': '''
                    SELECT country_long as "Country", COUNT(*) as "NPP Count"
                    FROM powerplants 
                    WHERE primary_fuel = 'Nuclear'
                    GROUP BY country_long 
                    ORDER BY COUNT(*) DESC 
                    LIMIT 10
                ''',
                'chart_type': 'pie',
                'business_question': 'Distribution of nuclear power plants by country'
            },
            {
                'name': '2. Top-10 countries by uranium production',
                'query': '''
                    SELECT country_long as "Country", SUM(production_tonnes_uranium) as "Uranium_Production_Tonnes"
                    FROM mines 
                    WHERE country_long IS NOT NULL
                    GROUP BY country_long 
                    ORDER BY SUM(production_tonnes_uranium) DESC 
                    LIMIT 10
                ''',
                'chart_type': 'bar',
                'business_question': 'Top-10 countries by uranium production'
            },
            {
                'name': '3. Average nuclear plant capacity by country',
                'query': '''
                    SELECT 
                        country_long as "Country", 
                        AVG(NULLIF(capacity_in_mw::numeric, 0)) as "Average Capacity of NPPs by MegaWatts"
                    FROM powerplants 
                    WHERE capacity_in_mw IS NOT NULL
                      AND NULLIF(TRIM(primary_fuel), '') IS NOT NULL
                      AND TRIM(primary_fuel) ILIKE '%nuclear%'
                    GROUP BY country_long 
                    ORDER BY "Average Capacity of NPPs by MegaWatts" DESC NULLS LAST
                    LIMIT 10
                ''',
                'chart_type': 'barh',
                'business_question': 'Average capacity of nuclear plants by country'
            },
            {
                'name': '4. Reserves and Production',
                'query': '''
                    SELECT r.country_long as "Country", 
                           r.tonnes_of_uranium as "Reserves_Tonnes",
                           COALESCE(m.production_tonnes_uranium, 0) as "Production_Tonnes"
                    FROM reserves r
                    LEFT JOIN mines m ON r.country_long = m.country_long
                    ORDER BY r.tonnes_of_uranium DESC
                    LIMIT 15
                ''',
                'chart_type': 'scatter',
                'business_question': 'Relationship between reserves and production by country'
            },
            {
                'name': '5. Distribution of Nuclear Plant Capacities',
                'query': '''
                    SELECT capacity_in_mw as "Capacity in MegaWatts"
                    FROM powerplants 
                    WHERE primary_fuel = 'Nuclear' AND capacity_in_mw IS NOT NULL AND capacity_in_mw > 0
                    ORDER BY capacity_in_mw
                ''',
                'chart_type': 'hist',
                'business_question': 'Distribution of Nuclear Plant Capacities'
            },
            {
                'name': '6. Uranium production top companies',
                'query': '''
                    SELECT 
                        c.company as "Company", 
                        c.tonnes_uranium as "Production_Tonnes"
                    FROM companies c
                    WHERE c.tonnes_uranium > 1000
                    ORDER BY c.tonnes_uranium DESC
                    LIMIT 20
                ''',
                'chart_type': 'line',
                'business_question': 'Uranium production by top companies'
            }
        ]
        
        interactive_queries = [
            {
                'name': 'Interactive Chart: Uranium Production Over Time',
                'query': '''
                    SELECT 
                        country_long as "Country",
                        "2007" as "Production_2007",
                        "2008" as "Production_2008", 
                        "2009" as "Production_2009",
                        "2010" as "Production_2010",
                        "2011" as "Production_2011",
                        "2012" as "Production_2012",
                        "2013" as "Production_2013",
                        "2014" as "Production_2014"
                    FROM production
                    WHERE country_long IS NOT NULL
                    ORDER BY "2014" DESC NULLS LAST
                    LIMIT 15
                ''',
                'business_question': 'Uranium Production Trends by Country'
            }
        ]
        
        excel_queries = [
            {
                'name': 'Nuclear power plants',
                'query': '''
                    SELECT p.name_of_powerplant as "Powerplant",
                           p.country_long as "Country",
                           p.capacity_in_mw as "Capacity_MW",
                           p.primary_fuel as "Primary_Fuel"
                    FROM powerplants p
                    WHERE p.primary_fuel = 'Nuclear'
                    ORDER BY p.capacity_in_mw DESC
                '''
            },
            {
                'name': 'Uranium companies',
                'query': '''
                    SELECT c.company as "Company",
                           c.tonnes_uranium as "Production_Tonnes"
                    FROM companies c
                    ORDER BY c.tonnes_uranium DESC
                '''
            },
            {
                'name': 'Uranium reserves',
                'query': '''
                    SELECT r.country_long as "Country",
                           r.tonnes_of_uranium as "Reserves_Tonnes"
                    FROM reserves r
                    ORDER BY r.tonnes_of_uranium DESC
                '''
            }
        ]
        
        print("\n" + "="*100)
        print("CREATING VISUALIZATIONS")
        print("="*100)
        
        for query_info in visualization_queries:
            print(f"\n{query_info['name']}")
            print("-" * 50)
            
            try:
                cur.execute(query_info['query'])
                rows = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
                
                if rows:
                    df = pd.DataFrame(rows, columns=columns)
                    
                    chart_path = create_charts(
                        df, 
                        query_info['name'].split('.')[0].strip(), 
                        query_info['chart_type'], 
                        query_info['business_question']
                    )
                    
                    print(f"\nChart data:")
                    print(df.to_string(index=False))
                    
                else:
                    print("No data available for chart creation")
                    
            except Exception as e:
                print(f"Error creating chart: {e}")
        
        print("\n" + "="*100)
        print("CREATING INTERACTIVE CHART")
        print("="*100)
        
        for query_info in interactive_queries:
            print(f"\n{query_info['name']}")
            print("-" * 50)
            
            try:
                cur.execute(query_info['query'])
                rows = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
                
                if rows:
                    df = pd.DataFrame(rows, columns=columns)
                    create_interactive_chart(df, "interactive", query_info['business_question'])
                else:
                    print("No data available for interactive chart")
                    
            except Exception as e:
                print(f"Error creating interactive chart: {e}")
        
        print("\n" + "="*100)
        print("EXPORT TO EXCEL")
        print("="*100)
        
        excel_data = {}
        for query_info in excel_queries:
            try:
                cur.execute(query_info['query'])
                rows = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
                
                if rows:
                    df = pd.DataFrame(rows, columns=columns)
                    excel_data[query_info['name']] = df
                    print(f"Prepared data for sheet: {query_info['name']} ({len(df)} rows)")
                else:
                    print(f"No data for sheet: {query_info['name']}")
                    
            except Exception as e:
                print(f"Error preparing data for Excel: {e}")
        
        if excel_data:
            try:
                excel_file = export_to_excel(excel_data, "uranium_industry_report.xlsx")
                print(f"Excel file created: {excel_file}")
            except Exception as e:
                print(f"Error creating Excel file: {e}")
        
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()
            print('Database connection closed.')


if __name__ == '__main__':
    run_queries()